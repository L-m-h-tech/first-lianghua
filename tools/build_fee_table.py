# -*- coding: utf-8 -*-
r"""把券商“投资者手续费率查询.xlsx”转换为回测运行时CSV。

仅维护转换工具使用 openpyxl；backtest.py 运行时只读标准库CSV，不新增运行依赖。
用法（在项目根目录）：
  D:\Python\python.exe tools\build_fee_table.py "C:\path\投资者手续费率查询.xlsx"
"""
from pathlib import Path
import argparse
import csv
import shutil
import sys
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
import config  # noqa: E402

# 各期货品种1手交易单位（合约乘数）：吨/克/千克/桶/立方米/指数点。
MULTIPLIERS = {
    # SHFE
    "RB": 10, "HC": 10, "SS": 5, "CU": 5, "AL": 5, "AO": 20, "ZN": 5, "PB": 5,
    "NI": 1, "SN": 1, "AU": 1000, "AG": 15, "RU": 10, "BR": 5, "FU": 10,
    "BU": 10, "SP": 10,
    # INE
    "SC": 1000, "NR": 10, "LU": 10, "BC": 5, "EC": 50,
    # DCE
    "A": 10, "B": 10, "M": 10, "Y": 10, "P": 10, "C": 10, "CS": 10, "RR": 10,
    "JD": 5, "LH": 16, "LG": 90, "L": 5, "V": 5, "PP": 5, "EG": 10, "EB": 5,
    "PG": 20, "J": 100, "JM": 60, "I": 100,
    # CZCE
    "SR": 10, "CF": 5, "CY": 5, "TA": 5, "MA": 10, "PX": 5, "PF": 5,
    "PR": 15, "SH": 30, "FG": 20, "SA": 20, "UR": 20, "RM": 10, "OI": 10,
    "PK": 5, "AP": 10, "CJ": 5, "SF": 5, "SM": 5,
    # GFEX
    "SI": 5, "LC": 1, "PS": 3,
}


def num(x):
    return 0.0 if x is None else float(x)


def convert(src: Path, out: Path, archive: bool = True):
    assert src.exists(), src
    cfg_syms = {m["sym"] for m in config.VARIETIES.values()}
    assert set(MULTIPLIERS) == cfg_syms, (len(MULTIPLIERS), len(cfg_syms))
    wb = load_workbook(src, data_only=True, read_only=True)
    ws = wb.active
    raw = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        sym = str(row[2] or "").strip().upper()
        if sym not in MULTIPLIERS:
            continue
        # 普通回测按投机账户；套保费率更低，不能让表格行顺序决定取值。
        if str(row[8] or "").strip() != "投机":
            continue
        assert sym not in raw, (sym, raw[sym], row)
        raw[sym] = row
    missing = sorted(set(MULTIPLIERS) - set(raw))
    assert not missing, missing

    out.parent.mkdir(parents=True, exist_ok=True)
    fields = ["sym", "name", "exchange", "account_flag", "multiplier",
              "open_amt_rate", "open_per_lot", "close_amt_rate", "close_per_lot",
              "today_amt_rate", "today_per_lot", "as_of"]
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for cname, meta in config.VARIETIES.items():
            sym = meta["sym"]
            row = raw[sym]
            w.writerow({
                "sym": sym,
                "name": cname,
                "exchange": row[1],
                "account_flag": row[8],
                "multiplier": MULTIPLIERS[sym],
                "open_amt_rate": num(row[9]),
                "open_per_lot": num(row[10]),
                "close_amt_rate": num(row[11]),
                "close_per_lot": num(row[12]),
                "today_amt_rate": num(row[13]),
                "today_per_lot": num(row[14]),
                "as_of": str(row[0]),
            })
    if archive:
        shutil.copy2(src, ROOT / "data" / src.name)
    return len(raw)


def main():
    parser = argparse.ArgumentParser(description="转换券商期货手续费Excel为backtest运行时CSV")
    parser.add_argument("src", help="投资者手续费率查询.xlsx路径")
    parser.add_argument("--out", default=str(ROOT / "data" / "futures_fees.csv"))
    parser.add_argument("--no-archive", action="store_true", help="不把原始xlsx复制到data目录")
    args = parser.parse_args()
    n = convert(Path(args.src), Path(args.out), not args.no_archive)
    print(f"wrote {args.out} rows={n}")


if __name__ == "__main__":
    main()
